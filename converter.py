from flask import Flask, request, jsonify
try:
    from flask_cors import CORS
except Exception:
    CORS = None
import openpyxl
import datetime
import re

app = Flask(__name__)
if CORS:
    CORS(app, resources={r"/convert": {"origins": "*"}})


def excel_serial_to_iso(n):
    try:
        num = float(n)
        base = datetime.datetime(1899, 12, 30)  # Excel 1900 system
        return (base + datetime.timedelta(days=num)).date().isoformat()
    except Exception:
        return "" if n is None else str(n)


def norm(s):
    s = "" if s is None else str(s)
    return s.replace("\uFEFF", "").strip()


def norm_header(h):
    h = norm(h)
    h = re.sub(r"\.+", "", h)
    h = re.sub(r"\s+", " ", h).strip().lower()
    return h


EXPECTED = {
    "sno", "s no", "serial", "name", "full name", "student name",
    "joining date", "joining dte", "joining", "date",
    "room no", "room number", "room no", "room",
    "room space", "space", "bed", "seat",
    "security", "securty", "security amount",
    "monthly rent", "rent", "per month", "rent amount",
    "rent date paid", "rent paid on", "paid date",
    "form fill charges", "fom fill charges", "form charges", "form fee",
    "educational institute", "institute", "university",
    "contact#", "contact #", "phone", "mobile", "contact",
    "floor", "ground floor", "first floor", "second floor",
    "any information", "information", "notes", "remarks", "comment"
}


CANON = {
    "name": "name", "full name": "name", "student name": "name",
    "joining date": "joiningDate", "joining dte": "joiningDate", "date": "joiningDate", "joining": "joiningDate",
    "room no": "roomNo", "room number": "roomNo", "room": "roomNo",
    "room space": "roomSpace", "space": "roomSpace", "bed": "roomSpace", "seat": "roomSpace",
    "security": "security", "securty": "security", "security amount": "security",
    "monthly rent": "monthlyRent", "rent": "monthlyRent", "per month": "monthlyRent", "rent amount": "monthlyRent",
    "rent date paid": "rentDatePaid", "rent paid on": "rentDatePaid", "paid date": "rentDatePaid",
    "form fill charges": "formFillCharges", "fom fill charges": "formFillCharges", "form charges": "formFillCharges", "form fee": "formFillCharges",
    "educational institute": "institute", "institute": "institute", "university": "institute",
    "contact#": "contact", "contact #": "contact", "phone": "contact", "mobile": "contact", "contact": "contact",
    "floor": "floor", "ground floor": "floor", "first floor": "floor", "second floor": "floor",
    "any information": "info", "information": "info", "notes": "info", "remarks": "info", "comment": "info"
}


def to_number(v):
    if v in (None, ""):
        return 0
    s = str(v).replace(",", "").strip()
    try:
        return int(s)
    except Exception:
        try:
            return float(s)
        except Exception:
            return 0


@app.post("/convert")
def convert():
    f = request.files.get("file")
    if not f:
        return jsonify({"error": "no file"}), 400

    wb = openpyxl.load_workbook(f, data_only=True)

    # choose best sheet
    best_ws, best_score = None, -1
    for ws in wb.worksheets:
        top = list(ws.iter_rows(values_only=True))[:10]
        s = 0
        for row in top:
            if not row:
                continue
            s = max(s, sum(1 for v in row if norm_header(v) in EXPECTED))
        if s > best_score:
            best_ws, best_score = ws, s
    ws = best_ws or wb.active

    rows = list(ws.iter_rows(values_only=True))

    # detect header row
    def header_score(r):
        if not r:
            return 0
        return sum(1 for v in r if norm_header(v) in EXPECTED)

    header_idx, best = -1, -1
    for i, r in enumerate(rows):
        s = header_score(r)
        if s > best:
            best, header_idx = s, i

    if header_idx < 0:
        return jsonify([])

    headers = [norm_header(h) for h in rows[header_idx]]
    out = []
    for r in rows[header_idx + 1:]:
        if not r or all(v in (None, "") for v in r):
            continue
        obj = {}
        for i, val in enumerate(r):
            if i >= len(headers):
                break
            k = CANON.get(headers[i])
            if not k:
                continue
            if k in ("joiningDate", "rentDatePaid"):
                obj[k] = excel_serial_to_iso(val) if isinstance(val, (int, float)) else norm(val)
            elif k in ("monthlyRent", "security", "formFillCharges"):
                obj[k] = to_number(val)
            else:
                obj[k] = norm(val)
        if any([obj.get("name"), obj.get("roomNo"), obj.get("contact"), obj.get("monthlyRent"), obj.get("security")]):
            out.append(obj)

    return jsonify(out)


if __name__ == "__main__":
    app.run(host="127.0.0.1", port=5001, debug=False)


